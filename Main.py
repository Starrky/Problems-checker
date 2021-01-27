import configs as cfgs
import selenium
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import glob
import os
import os.path
from os import path
import time
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join
from itertools import chain

""" SPECIFY SHEETNAMES HERE """
countries_folder = cfgs.Countries_folder
onlyfiles = [f for f in listdir(countries_folder) if isfile(join(countries_folder, f))]

# Sheetnames with POS/ BOS data in excel files 
sheets = cfgs.zbx_sheets

# Paths and links setup
problems_link = cfgs.zbx_problems_link
zbx = cfgs.zbx_link
username = cfgs.zbx_username
password = cfgs.zbx_password

download_loc = cfgs.download_loc
countries_folder = cfgs.Countries_folder
filename = 'zbx_problems_export.csv'
report_file = f'{download_loc}{filename}'


# Delete old zbx problems export file
files = glob.glob(f"{download_loc}*")
for f in files:
    os.remove(f)


# Webdriver setup
options = webdriver.ChromeOptions()
options.add_argument('ignore-certificate-errors')
options.add_argument("--headless")
prefs = {"download.default_directory": download_loc}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(
    ChromeDriverManager().install(), options=options)


# Fetching file
driver.get(problems_link)
driver.find_element_by_id("login").click()
driver.find_element_by_id("name").send_keys(username)
driver.find_element_by_id("password").send_keys(password)
driver.find_element_by_id("enter").click()
driver.find_element_by_xpath(
    "/html/body/main/div[1]/div[2]/nav/form/ul/li[1]/button").click()


# Check if file was downloaded and close driver
time.sleep(5)
while path.exists(report_file) is False:
    continue

else:
    driver.close()


# Process the file and make dictionaries from hosts with issues
BOS = []
POS1 = []
POS2 = []
POS3 = []
POS4 = []
mapping = {}

with open(report_file) as fo:
    reader = csv.DictReader(fo)
    for row in reader:
        mapping[row["Host"]] = row["Problem"]

for key, value in mapping.items():
    if value == "BOS is unavailable by ICMP":
        BOS.append(key)

    if value == "POS1 is unavailable by ICMP":
        POS1.append(key)

    if value == "POS2 is unavailable by ICMP":
        POS2.append(key)

    if value == "POS3 is unavailable by ICMP":
        POS3.append(key)

    if value == "POS4 is unavailable by ICMP":
        POS4.append(key)


print(f'\n\n *** OFFLINE DEVICES SUMMARY ***\n\nPOS1 offline for: {BOS}\n\nPOS2 offline for: {POS1}'
      f'\n\nPOS3 offline for: {POS2}\n\nPOS4 offline for: {POS3}\n\nBOS offline for: {POS4}')

# Read country file and change columns showing status of devices
for file in onlyfiles:
    wb = load_workbook(f'{cfgs.Countries_folder}{file}')
    worksheets = wb.sheetnames
    for sheet in worksheets:
        if sheet in sheets:
            ws = wb[sheet]
            # print(file)
        else:
            pass

# Iterate over all rows in excel file
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            store = str(cell.value).zfill(4)
            BOS_name = ws[f'E{cell.row}'].value
            POS1_name = ws[f'G{cell.row}'].value
            POS2_name = ws[f'I{cell.row}'].value
            POS3_name = ws[f'K{cell.row}'].value
            POS4_name = ws[f'M{cell.row}'].value

        # Change all to OK
            ws[f'F{cell.row}'] = 'ok'
            ws[f'H{cell.row}'] = 'ok'
            ws[f'J{cell.row}'] = 'ok'
            ws[f'L{cell.row}'] = 'ok'
            ws[f'N{cell.row}'] = 'ok'

        # Check dictionaries
            if store in BOS and BOS_name != 'brak':
                ws[f'F{cell.row}'] = 'failed'
                # print(f'BOS offline for {store}')
            
            elif store in POS1 and POS1_name != 'brak':
                ws[f'H{cell.row}'] = 'failed'
                # print(f'POS1 offline for {store}')

            elif store in POS2 and POS2_name != 'brak':
                ws[f'J{cell.row}'] = 'failed'
                # print(f'POS2 offline for {store}')

            elif store in POS3 and  POS3_name != 'brak':
                ws[f'L{cell.row}'] = 'failed'
                # print(f'POS3 offline for {store}')

            elif store in POS4 and POS4_name != 'brak':
                ws[f'N{cell.row}'] = 'failed'
                # print(f'POS4 offline for {store}')

        # Braki
            elif POS1_name == 'brak':
                ws[f'H{cell.row}'] = 'brak'
                # print(f'POS4 offline for {store}')
           
            elif POS2_name == 'brak':
                ws[f'J{cell.row}'] = 'brak'
                # print(f'POS4 offline for {store}')
            
            elif POS3_name == 'brak':
                ws[f'L{cell.row}'] = 'brak'
                # print(f'POS4 offline for {store}')

            elif POS4_name == 'brak':
                ws[f'N{cell.row}'] = 'brak'
                # print(f'POS4 offline for {store}')
            
    # save file 
    wb.save(f'{cfgs.Countries_folder}{file}')
